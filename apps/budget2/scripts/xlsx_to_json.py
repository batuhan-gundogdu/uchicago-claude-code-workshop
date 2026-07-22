"""One-time converter: budget_fake.xlsx -> budget.json.

Reads the six sheets of the family budget workbook with openpyxl
(data_only=True, no formulas in the output) and produces a dict matching
the app's budget.json data contract.
"""

import calendar
from datetime import date, datetime

import openpyxl

# Month header (as it appears in the Monthly Budget / Groceries / Enjoy
# sheets) -> (year, month number). Nov/Dec belong to 2025, Jan-Jun to 2026.
MONTH_MAP = {
    "November": (2025, 11),
    "December": (2025, 12),
    "January": (2026, 1),
    "February": (2026, 2),
    "March": (2026, 3),
    "April": (2026, 4),
    "May": (2026, 5),
    "June": (2026, 6),
}

# The 17 required line ids, in a fixed order, with their category/type.
LINE_DEFS = [
    ("mortgage", "expense", "Housing"),
    ("hoa", "expense", "Housing"),
    ("internet", "expense", "Utilities"),
    ("comed", "expense", "Utilities"),
    ("water", "expense", "Utilities"),
    ("gas", "expense", "Utilities"),
    ("groceries", "expense", "Food"),
    ("car_loan", "expense", "Transportation"),
    ("car_expenses", "expense", "Transportation"),
    ("home_expenses", "expense", "Home"),
    ("enjoy", "expense", "Lifestyle"),
    ("vacation_savings", "expense", "Savings"),
    ("rainy_day_savings", "expense", "Savings"),
    ("payment_to_principal", "expense", "Debt"),
    ("rent_income", "income", "Income"),
    ("batu_contribution", "income", "Income"),
    ("natalie_contribution", "income", "Income"),
]

DETAILED_LINES = {"groceries", "enjoy", "home_expenses", "car_expenses"}
FIXED_LINES = {"mortgage", "hoa", "internet", "car_loan", "batu_contribution", "natalie_contribution"}

LINE_NAMES = {
    "mortgage": "Mortgage",
    "hoa": "HOA",
    "internet": "Internet",
    "comed": "Comed",
    "water": "Water",
    "gas": "Gas",
    "groceries": "Groceries",
    "car_loan": "Car Loan",
    "car_expenses": "Car Expenses",
    "home_expenses": "Home Expenses",
    "enjoy": "Enjoy (order + eat out)",
    "vacation_savings": "Vacation Savings",
    "rainy_day_savings": "Rainy Day Savings",
    "payment_to_principal": "Payment to Principal",
    "rent_income": "Rent Income",
    "batu_contribution": "Batu Contribution",
    "natalie_contribution": "Natalie Contribution",
}


def _iso(year, month, day):
    """Clamp day into the valid range for year/month and return an ISO date string."""
    last_day = calendar.monthrange(year, month)[1]
    day = max(1, min(day, last_day))
    return date(year, month, day).isoformat()


def _to_num(v):
    """Return v as a float if it is a real number, else None (skips text like the stray '`')."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return None


def _parse_monthly_budget(ws):
    """Parse the 'Monthly Budget' sheet.

    Returns (lines, monthly_actuals) where lines is a dict line_id -> {anticipated, recurringDefault}
    and monthly_actuals is a dict "YYYY-MM" -> {line_id: amount}.
    """
    header = [c.value for c in ws[1]]
    # Month header cells start at column C (index 2, 0-based) through the 8th month.
    month_cols = []  # list of (col_index_0based, (year, month))
    for idx, name in enumerate(header):
        if name in MONTH_MAP:
            month_cols.append((idx, MONTH_MAP[name]))

    lines = {}
    monthly_actuals = {}
    for ym in month_cols:
        year, month = ym[1]
        key = f"{year}-{month:02d}"
        monthly_actuals.setdefault(key, {})

    # Raw row values keyed by label, for folding the three car rows.
    raw = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        label = row[0]
        if not label:
            continue
        anticipated = _to_num(row[1])
        months = {}
        for idx, (year, month) in month_cols:
            val = _to_num(row[idx])
            if val is not None:
                months[f"{year}-{month:02d}"] = val
        raw[label] = {"anticipated": anticipated, "months": months}

    def set_line(line_id, anticipated, is_fixed):
        lines[line_id] = {
            "anticipated": anticipated,
            "recurringDefault": anticipated if is_fixed else None,
        }

    # Simple 1:1 rows.
    simple = {
        "Mortgage": "mortgage",
        "HOA": "hoa",
        "Internet": "internet",
        "Comed": "comed",
        "Water": "water",
        "Gas": "gas",
        "Groceries": "groceries",
        "Enjoy (order + eat out)": "enjoy",
        "Home Expenses": "home_expenses",
        "Vacation Saving": "vacation_savings",
        "Rainy Day Savings": "rainy_day_savings",
        "Payment to Principal": "payment_to_principal",
        "Rent Income": "rent_income",
        "Batu Contribution": "batu_contribution",
        "Natalie Contribution": "natalie_contribution",
    }
    # Lines whose per-month values should be recorded into monthlyActuals.
    variable_actual_lines = {
        "Comed": "comed",
        "Water": "water",
        "Gas": "gas",
        "Rent Income": "rent_income",
        "Vacation Saving": "vacation_savings",
        "Rainy Day Savings": "rainy_day_savings",
        "Payment to Principal": "payment_to_principal",
    }

    for label, line_id in simple.items():
        r = raw.get(label)
        if r is None:
            continue
        is_fixed = line_id in FIXED_LINES
        set_line(line_id, r["anticipated"], is_fixed)

    for label, line_id in variable_actual_lines.items():
        r = raw.get(label)
        if r is None:
            continue
        for key, val in r["months"].items():
            monthly_actuals[key][line_id] = val

    # Fold the three car rows: Car -> car_loan (fixed); Car Insurance +
    # Gas + Park for Car -> car_expenses (variable, detailed).
    car = raw.get("Car", {"anticipated": None, "months": {}})
    car_ins = raw.get("Car Insurance", {"anticipated": None, "months": {}})
    car_gas = raw.get("Gas + Park for Car", {"anticipated": None, "months": {}})

    set_line("car_loan", car["anticipated"], True)
    for key, val in car["months"].items():
        monthly_actuals[key]["car_loan"] = val

    car_exp_anticipated = (car_ins["anticipated"] or 0) + (car_gas["anticipated"] or 0)
    set_line("car_expenses", car_exp_anticipated, False)
    all_keys = set(car_ins["months"]) | set(car_gas["months"])
    for key in all_keys:
        monthly_actuals[key]["car_expenses"] = car_ins["months"].get(key, 0) + car_gas["months"].get(key, 0)

    return lines, monthly_actuals


def _grocery_like_entries(ws, line_id, include_store):
    """Parse a Groceries/Enjoy-shaped sheet: a 'Month' header row, a TOTAL row,
    then a leftmost day-serial column with numeric cells under each month
    column. Every numeric cell under a month column becomes one entry - we
    capture all transactions rather than trying to match the spreadsheet's
    own (sometimes incomplete) cached TOTAL.
    """
    # Find the header row (contains "Month" in column A) and the row above
    # it that may carry store names (for Groceries only).
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Month":
            header_row = r
            break
    if header_row is None:
        raise ValueError("Could not find 'Month' header row")

    first_data_row = header_row + 2  # skip the TOTAL row right below the header

    month_cols = []  # (col_index_1based, (year, month))
    for c in range(2, ws.max_column + 1):
        name = ws.cell(row=header_row, column=c).value
        if name in MONTH_MAP:
            month_cols.append((c, MONTH_MAP[name]))

    entries = []
    for c, (year, month) in month_cols:
        store = None
        if include_store:
            name1 = ws.cell(row=header_row - 2, column=c).value
            name2 = ws.cell(row=header_row - 1, column=c).value
            names = [n for n in (name1, name2) if n]
            store = " / ".join(names) if names else None

        for r in range(first_data_row, ws.max_row + 1):
            day_serial = _to_num(ws.cell(row=r, column=1).value)
            amount = _to_num(ws.cell(row=r, column=c).value)
            if day_serial is None or amount is None:
                continue
            day = int(round(day_serial))
            entries.append({
                "id": f"{line_id}-{year}{month:02d}-{r}",
                "date": _iso(year, month, day),
                "line": line_id,
                "amount": amount,
                "item": None,
                "store": store,
                "note": None,
            })
    return entries


def _parse_home_expenses(ws):
    """Parse Date/Item/Price triples repeated across month blocks."""
    entries = []
    counter = 0
    for block_start in range(1, ws.max_column + 1, 3):
        date_col, item_col, price_col = block_start, block_start + 1, block_start + 2
        if price_col > ws.max_column:
            break
        for r in range(4, ws.max_row + 1):
            d = ws.cell(row=r, column=date_col).value
            item = ws.cell(row=r, column=item_col).value
            price = _to_num(ws.cell(row=r, column=price_col).value)
            if not isinstance(d, datetime) or price is None:
                continue
            counter += 1
            entries.append({
                "id": f"home_expenses-{counter}",
                "date": d.date().isoformat(),
                "line": "home_expenses",
                "amount": price,
                "item": item,
                "store": None,
                "note": None,
            })
    return entries


# Bookkeeping skip rule (positive-only): the workbook's monthly auto-transfer
# into the Vacation / Rainy Day funds (the recurring ~470-545 contributions
# that are already captured via monthlyActuals["vacation_savings"/
# "rainy_day_savings"]) also shows up as rows in the Bookkeeping ledger,
# often sharing a date with a genuine one-off (a car/student loan payment
# into the house fund). We skip a Vacation/Rainy Day cell only when it is a
# positive amount in that narrow recurring band - a negative amount in the
# same range is a genuine one-off (e.g. a real withdrawal) and is never
# skipped. House (Arlington) column cells are always kept, since every
# Arlington movement here is a genuine one-off (sale, loan, principal
# payment, or a car/student payment funneled into the house account).
RECURRING_SAVINGS_MIN = 450
RECURRING_SAVINGS_MAX = 550


def _parse_bookkeeping(ws):
    ledger = []
    counter = 0
    fund_cols = {2: "vacation", 3: "rainy_day", 4: "house"}
    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        if not isinstance(d, datetime):
            continue
        note = ws.cell(row=r, column=5).value
        for col, fund in fund_cols.items():
            amount = _to_num(ws.cell(row=r, column=col).value)
            if amount is None:
                continue
            if fund != "house" and RECURRING_SAVINGS_MIN <= amount <= RECURRING_SAVINGS_MAX:
                continue  # skip: duplicates the auto-fed monthly savings contribution (positive-only)
            counter += 1
            ledger.append({
                "id": f"fund-{counter}",
                "date": d.date().isoformat(),
                "fund": fund,
                "amount": amount,
                "note": note,
            })
    return ledger


def build_budget(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    line_data, monthly_actuals = _parse_monthly_budget(wb["Monthly Budget"])

    groceries_entries = _grocery_like_entries(wb["Groceries"], "groceries", include_store=True)
    enjoy_entries = _grocery_like_entries(wb["Enjoy"], "enjoy", include_store=False)
    home_entries = _parse_home_expenses(wb["Home Expenses"])
    entries = groceries_entries + enjoy_entries + home_entries

    funds_ledger = _parse_bookkeeping(wb["Bookkeeping"])

    lines = []
    for line_id, line_type, category in LINE_DEFS:
        d = line_data.get(line_id, {"anticipated": None, "recurringDefault": None})
        lines.append({
            "id": line_id,
            "name": LINE_NAMES[line_id],
            "type": line_type,
            "category": category,
            "detailed": line_id in DETAILED_LINES,
            "anticipated": d["anticipated"],
            "recurringDefault": d["recurringDefault"],
        })

    stores = sorted({e["store"] for e in groceries_entries if e["store"]})

    return {
        "version": 1,
        "settings": {
            "stores": stores,
            "lines": lines,
        },
        "entries": entries,
        "monthlyActuals": monthly_actuals,
        "fundsLedger": funds_ledger,
    }
